import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import uuid
import os

# Загружаем файл экселя
wb = load_workbook('Координаты пунктов охранных зон пунктов ГГC.xlsx')


# печатаем список листов если надо
sheets = wb.sheetnames
# for sheet in sheets:
#     print(sheet)

# Делаем первый лист активным
sheet = wb.active
# Создаем вложенный список, где каждый элемент это строка из эксель таблицы
total_sp = []
# 2 - это со 2 строки и ниже со 2 столбца
for i in range(2, sheet.max_row):
    sp = []
    for column in sheet.iter_cols(2, sheet.max_column):
        if column[i].value == None:
            continue
        else:
            sp.append(column[i].value)
    total_sp.append(sp)

# получаем другой лист если есть
if len(sheets) > 1:
    sheet2 = wb['на здании ГГС']
    # Делаем второй лист активным
    wb.active = 1
    for i in range(2, sheet2.max_row):
        sp = []
        for column in sheet2.iter_cols(2, sheet2.max_column):
            if column[i].value == None:
                continue
            else:
                sp.append(column[i].value)
        total_sp.append(sp)

# Удаляем последний элемент (он пустой список), если надо
# del total_sp[-1]
# удаляем NONE
for nn in total_sp:
    if nn == [] or nn[0] == None:
        indx = total_sp.index(nn)
        del total_sp[indx]
# Добавляем замыкающую точку в конце списка
for gg in total_sp:
    gg.append(gg[3])
    gg.append(gg[4])
# for vv in total_sp:
#     print(vv)



# Добавляем пространство имен, если есть
ET.register_namespace("", "urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8")
ET.register_namespace("", "urn://x-artefacts-rosreestr-ru/incoming/territory-to-gkn/1.0.4")
ET.register_namespace("p1", "http://www.w3.org/2001/XMLSchema-instance")
ET.register_namespace("Spa2", "urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1")
ET.register_namespace("CadEng4", "urn://x-artefacts-rosreestr-ru/commons/complex-types/cadastral-engineer/4.1.1")
ET.register_namespace("Doc5", "urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1")
ET.register_namespace("schemaLocation", "urn://x-artefacts-rosreestr-ru/incoming/territory-to-gkn/1.0.4 TerritoryToGKN_v01.xsd")
ET.register_namespace("tns", "urn://x-artefacts-smev-gov-ru/supplementary/commons/1.0.1")
ET.register_namespace("DocI5", "urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1")
ET.register_namespace("Sen5", "urn://x-artefacts-rosreestr-ru/commons/complex-types/sender/5.0.1")
ET.register_namespace("Org4", "urn://x-artefacts-rosreestr-ru/commons/complex-types/organization/4.0.1")
ET.register_namespace("AdrInp6", "urn://x-artefacts-rosreestr-ru/commons/complex-types/address-input/6.0.1")
ET.register_namespace("Gov5", "urn://x-artefacts-rosreestr-ru/commons/complex-types/governance/5.0.1")
ET.register_namespace("Person5", "urn://x-artefacts-rosreestr-ru/commons/complex-types/person/5.0.2")
ET.register_namespace("Zon4", "urn://x-artefacts-rosreestr-ru/commons/complex-types/zone/4.2.2")
ET.register_namespace("schemaLocation", "urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8 ZoneToGKN_v05.xsd")

# Парсим хмл с территориями и зонами
tree = ET.parse('Территория.xml')
root = tree.getroot()
tree_1 = ET.parse('Зона.xml')
root_1 = tree_1.getroot()



# Проверяет, существует папка "Готовые", если нет, то создает, иначе пропускает
if not os.path.isdir("Готовые"):
     os.mkdir("Готовые")





# Ввод даты и номера приказа/ов
count_prikaz = input("Приказов больше одного? Введите да или нет ")
if count_prikaz == 'нет':
    cod_documenta = input("Введите код документа(приказа): ")
    name_prikaz = input("Введите имя документа(приказа): ")
    number_prikaz = input("Введите номер документа(приказа): ")
    date_prikaz = input("Введите дату документа(приказа): ")
    publication = input("Подлежит ли документ(приказ) публикации? Введите да или нет: ")
    # Проходимся по общему списку и в каждом меняем координаты, при необходимости добавляем новые
    for i in range(len(total_sp)):
        ychastok = + i
        raion = total_sp[ychastok][2]
        zona = total_sp[ychastok][4][0]
        x = 3
        y = 4
        count = 1
        number_tochki = 6
        number_coord = 5
        # Если 4 точки
        if len(total_sp[ychastok][12:]) == 1:
            for neighbor in root.iter(
                    '{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate'):
                neighbor.attrib['X'] = total_sp[ychastok][x]
                neighbor.attrib['Y'] = total_sp[ychastok][y]
                x += 2
                y += 2
            # Генерируем рандомный GUID
            guid = uuid.uuid4()
            # Конвертируем в строку
            str_guid = str(guid)
            root.attrib['GUID'] = str_guid
            root[0][1][0][1].text = "1047727043561"
            guid_1 = uuid.uuid4()
            str_guid_1 = str(guid_1)
            root[2][0].attrib["Name"] = f'МСК-50, зона {zona}'
            root_1.attrib['GUID'] = str_guid_1
            root_1[3][0][4][1][1].attrib['GUID'] = str_guid
            root_1[3][0][4][1][1].attrib['Name'] = f'TerritoryToGKN_{str_guid}.xml'
            root_1[0][1].text = f'ZoneToGKN_{str_guid_1}'
            root_1[0][3].text = date_prikaz
            root_1[2][0][0].text = cod_documenta
            root_1[2][0][1].text = name_prikaz
            root_1[2][0][2].text = number_prikaz
            root_1[2][0][3].text = date_prikaz
            if publication == "нет":
                root_1[2][0][5].tag = '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication'
                root_1[2][0][6].attrib[
                    'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
            else:
                root_1[2][0].remove(root_1[2][0][5])
                root_1[2][0][5].attrib[
                    'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
            root_1[3][0][0].text = f'{total_sp[ychastok][2]}'
            root_1[3][0][2].text = f'{total_sp[ychastok][1]}'
            if not os.path.isdir(f'.//Готовые/ZoneToGKN_{str_guid_1}'):
                os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}')
            os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}/Applied_files')
            tree.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//TerritoryToGKN_{str_guid}.xml',
                       encoding='utf-8', xml_declaration=True)
            tree_1.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//ZoneToGKN_{str_guid_1}.xml',
                         encoding='utf-8', xml_declaration=True)
            tree = ET.parse('Территория.xml')
            root = tree.getroot()
            tree_1 = ET.parse('Зона.xml')
            root_1 = tree_1.getroot()
        # Если больше 4 точек
        if len(total_sp[ychastok][12:]) > 1:
            for neighbor_1 in root.iter(
                    '{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate'):
                DeltaGeopoint = str(neighbor_1.attrib['DeltaGeopoint'])
                GeopointOpred = str(neighbor_1.attrib['GeopointOpred'])
                attrib_1 = {"TypeUnit": "Точка", "SuNmb": str(number_tochki)}
                attrib_2 = {"X": str(total_sp[ychastok][x]), "Y": str(total_sp[ychastok][y]),
                            "NumGeopoint": str(number_tochki),
                            "DeltaGeopoint": DeltaGeopoint,
                            "GeopointOpred": GeopointOpred}
                neighbor_1.attrib['X'] = total_sp[ychastok][x]
                neighbor_1.attrib['Y'] = total_sp[ychastok][y]
                x += 2
                y += 2
                count += 1
                if count == 6:  # шестерка потому что, после пятой точки , становится 6 и тогда заходит в этот if.
                    root[1][0][4].attrib['SuNmb'] = '5'
                    neighbor_1.attrib['NumGeopoint'] = "5"
                    attrib_2['X'] = str(total_sp[ychastok][x])
                    attrib_2['Y'] = str(total_sp[ychastok][y])
                    count += 1
                    x += 2
                    y += 2
                if count > 6:
                    konec = len(total_sp[i][13:])
                    while konec != 2:
                        ET.SubElement(root[1][0],
                                      "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}SpelementUnit",
                                      attrib_1)
                        ET.SubElement(root[1][0][number_coord],
                                      "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate ",
                                      attrib_2)
                        number_tochki += 1
                        attrib_1["SuNmb"] = str(number_tochki)
                        number_coord += 1
                        attrib_2['X'] = str(total_sp[ychastok][x])
                        attrib_2['Y'] = str(total_sp[ychastok][y])
                        attrib_2['NumGeopoint'] = str(number_tochki)
                        x += 2
                        y += 2
                        konec -= 2
                    attrib_1["SuNmb"] = str(1)
                    attrib_2['NumGeopoint'] = str(1)
                    ET.SubElement(root[1][0],
                                  "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}SpelementUnit",
                                  attrib_1)
                    ET.SubElement(root[1][0][number_coord],
                                  "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate ",
                                  attrib_2)
            # Генерируем рандомный GUID
            guid = uuid.uuid4()
            # Конвертируем в строку
            str_guid = str(guid)
            root.attrib['GUID'] = str_guid
            root[0][1][0][1].text = "1047727043561"
            guid_1 = uuid.uuid4()
            str_guid_1 = str(guid_1)
            root[2][0].attrib["Name"] = f'МСК-50, зона {zona}'
            root_1.attrib['GUID'] = str_guid_1
            root_1[3][0][4][1][1].attrib['GUID'] = str_guid
            root_1[3][0][4][1][1].attrib['Name'] = f'TerritoryToGKN_{str_guid}.xml'
            root_1[0][1].text = f'ZoneToGKN_{str_guid_1}'
            root_1[0][3].text = date_prikaz
            root_1[2][0][0].text = cod_documenta
            root_1[2][0][1].text = name_prikaz
            root_1[2][0][2].text = number_prikaz
            root_1[2][0][3].text = date_prikaz
            if publication == "нет":
                root_1[2][0][5].tag = '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication'
                root_1[2][0][6].attrib[
                    'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
            else:
                root_1[2][0].remove(root_1[2][0][5])
                root_1[2][0][5].attrib[
                    'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
            root_1[3][0][0].text = f'{total_sp[ychastok][2]}'
            root_1[3][0][2].text = f'{total_sp[ychastok][1]}'
            if not os.path.isdir(f'.//Готовые/ZoneToGKN_{str_guid_1}'):
                os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}')
            os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}/Applied_files')
            tree.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//TerritoryToGKN_{str_guid}.xml',
                       encoding='utf-8', xml_declaration=True)
            tree_1.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//ZoneToGKN_{str_guid_1}.xml',
                         encoding='utf-8', xml_declaration=True)

            tree = ET.parse('Территория.xml')
            root = tree.getroot()
            tree_1 = ET.parse('Зона.xml')
            root_1 = tree_1.getroot()

if count_prikaz == "да":
    kolvo_prikaz = int(input("Сколько приказов добавить в xml? Введите число: "))
    if kolvo_prikaz == 2:
        cod_documenta = input("Введите код первого документа(приказа): ")
        name_prikaz = input("Введите имя первого документа(приказа): ")
        number_prikaz = input("Введите номер первого документа(приказа): ")
        date_prikaz = input("Введите дату первого документа(приказа): ")
        publication = input("Подлежит ли документ(приказ) публикации? Введите да или нет: ")
        cod_documenta_1 = input("Введите код второго документа(приказа): ")
        name_prikaz_1 = input("Введите имя второго документа(приказа): ")
        number_prikaz_1 = input("Введите номер второго документа(приказа): ")
        date_prikaz_1 = input("Введите дату второго документа(приказа): ")
        publication_1 = input("Подлежит ли документ(приказ) публикации? Введите да или нет: ")

        for i in range(len(total_sp)):
            ychastok = + i
            raion = total_sp[ychastok][2]
            zona = total_sp[ychastok][4][0]
            x = 3
            y = 4
            count = 1
            number_tochki = 6
            number_coord = 5
            # Условие, где 4 точки
            if len(total_sp[ychastok][12:]) == 1:
                for neighbor in root.iter(
                        '{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate'):
                    neighbor.attrib['X'] = total_sp[ychastok][x]
                    neighbor.attrib['Y'] = total_sp[ychastok][y]
                    x += 2
                    y += 2
                # Генерируем рандомный GUID
                guid = uuid.uuid4()
                # Конвертируем в строку
                str_guid = str(guid)
                root.attrib['GUID'] = str_guid
                root[0][1][0][1].text = "1047727043561"
                guid_1 = uuid.uuid4()
                str_guid_1 = str(guid_1)
                root[2][0].attrib["Name"] = f'МСК-50, зона {zona}'
                root_1.attrib['GUID'] = str_guid_1
                root_1[3][0][4][1][1].attrib['GUID'] = str_guid
                root_1[3][0][4][1][1].attrib['Name'] = f'TerritoryToGKN_{str_guid}.xml'
                root_1[0][1].text = f'ZoneToGKN_{str_guid_1}'
                root_1[2][0][0].text = cod_documenta
                root_1[0][3].text = date_prikaz
                root_1[2][0][1].text = name_prikaz
                root_1[2][0][2].text = number_prikaz
                root_1[2][0][3].text = date_prikaz
                if publication == "нет":
                    root_1[2][0][
                        5].tag = '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication'
                    root_1[2][0][6].attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
                else:
                    root_1[2][0].remove(root_1[2][0][5])
                    root_1[2][0][5].attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
                root_1[3][0][0].text = f'{total_sp[ychastok][2]}'
                root_1[3][0][2].text = f'{total_sp[ychastok][1]}'
                # Добавляем новый подэлемент
                ET.SubElement(root_1[2], '{urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8}Document')
                # Добавляем в него еще подэлементы с новым текстом
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}CodeDocument').text = cod_documenta_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Name').text = name_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Number').text = number_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Date').text = date_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}IssueOrgan').text = 'Управление Росреестра по Московской области'
                if publication_1 == 'нет':
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication').text = 'Решение не подлежит публикации в связи с отсутствием необходимости опубликования таких решений в соответствии с Постановлением Правительства РФ от 21 августа 2019 г. N 1080 "Об охранных зонах пунктов государственной геодезической сети, государственной нивелирной сети и государственной гравиметрической сети"'
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_1}_№_{number_prikaz_1}.pdf'
                    root_1[2][1][6].attrib['Kind'] = '01'
                else:
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_1}_№_{number_prikaz_1}.pdf'
                    root_1[2][1][5].attrib['Kind'] = '01'
                if not os.path.isdir(f'.//Готовые/ZoneToGKN_{str_guid_1}'):
                    os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}')
                os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}/Applied_files')
                tree.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//TerritoryToGKN_{str_guid}.xml',
                           encoding='utf-8', xml_declaration=True)
                tree_1.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//ZoneToGKN_{str_guid_1}.xml',
                             encoding='utf-8', xml_declaration=True)
                tree = ET.parse('Территория.xml')
                root = tree.getroot()
                tree_1 = ET.parse('Зона.xml')
                root_1 = tree_1.getroot()
            # Условие, где  больше 4 точек
            if len(total_sp[ychastok][12:]) > 1:
                for neighbor_1 in root.iter(
                        '{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate'):
                    DeltaGeopoint = str(neighbor_1.attrib['DeltaGeopoint'])
                    GeopointOpred = str(neighbor_1.attrib['GeopointOpred'])
                    attrib_1 = {"TypeUnit": "Точка", "SuNmb": str(number_tochki)}
                    attrib_2 = {"X": str(total_sp[ychastok][x]), "Y": str(total_sp[ychastok][y]),
                                "NumGeopoint": str(number_tochki),
                                "DeltaGeopoint": DeltaGeopoint,
                                "GeopointOpred": GeopointOpred}
                    neighbor_1.attrib['X'] = total_sp[ychastok][x]
                    neighbor_1.attrib['Y'] = total_sp[ychastok][y]
                    x += 2
                    y += 2
                    count += 1
                    if count == 6:  # шестерка потому что, после пятой точки , становится 6 и тогда заходит в этот if.
                        root[1][0][4].attrib['SuNmb'] = '5'
                        neighbor_1.attrib['NumGeopoint'] = "5"
                        attrib_2['X'] = str(total_sp[ychastok][x])
                        attrib_2['Y'] = str(total_sp[ychastok][y])
                        count += 1
                        x += 2
                        y += 2
                    if count > 6:
                        konec = len(total_sp[i][13:])
                        while konec != 2:
                            ET.SubElement(root[1][0],
                                          "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}SpelementUnit",
                                          attrib_1)
                            ET.SubElement(root[1][0][number_coord],
                                          "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate ",
                                          attrib_2)
                            number_tochki += 1
                            attrib_1["SuNmb"] = str(number_tochki)
                            number_coord += 1
                            attrib_2['X'] = str(total_sp[ychastok][x])
                            attrib_2['Y'] = str(total_sp[ychastok][y])
                            attrib_2['NumGeopoint'] = str(number_tochki)
                            x += 2
                            y += 2
                            konec -= 2
                        attrib_1["SuNmb"] = str(1)
                        attrib_2['NumGeopoint'] = str(1)
                        ET.SubElement(root[1][0],
                                      "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}SpelementUnit",
                                      attrib_1)
                        ET.SubElement(root[1][0][number_coord],
                                      "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate ",
                                      attrib_2)
                # Генерируем рандомный GUID
                guid = uuid.uuid4()
                # Конвертируем в строку
                str_guid = str(guid)
                root.attrib['GUID'] = str_guid
                root[0][1][0][1].text = "1047727043561"
                guid_1 = uuid.uuid4()
                str_guid_1 = str(guid_1)
                root[2][0].attrib["Name"] = f'МСК-50, зона {zona}'
                root_1.attrib['GUID'] = str_guid_1
                root_1[3][0][4][1][1].attrib['GUID'] = str_guid
                root_1[3][0][4][1][1].attrib['Name'] = f'TerritoryToGKN_{str_guid}.xml'
                root_1[0][1].text = f'ZoneToGKN_{str_guid_1}'
                root_1[2][0][0].text = cod_documenta
                root_1[0][3].text = date_prikaz
                root_1[2][0][1].text = name_prikaz
                root_1[2][0][2].text = number_prikaz
                root_1[2][0][3].text = date_prikaz
                if publication == "нет":
                    root_1[2][0][
                        5].tag = '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication'
                    root_1[2][0][6].attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
                else:
                    root_1[2][0].remove(root_1[2][0][5])
                    root_1[2][0][5].attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
                root_1[3][0][0].text = f'{total_sp[ychastok][2]}'
                root_1[3][0][2].text = f'{total_sp[ychastok][1]}'
                # Добавляем новый подэлемент
                ET.SubElement(root_1[2], '{urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8}Document')
                # Добавляем в него еще подэлементы с новым текстом
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}CodeDocument').text = cod_documenta_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Name').text = name_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Number').text = number_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Date').text = date_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}IssueOrgan').text = 'Управление Росреестра по Московской области'
                if publication_1 == 'нет':
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication').text = 'Решение не подлежит публикации в связи с отсутствием необходимости опубликования таких решений в соответствии с Постановлением Правительства РФ от 21 августа 2019 г. N 1080 "Об охранных зонах пунктов государственной геодезической сети, государственной нивелирной сети и государственной гравиметрической сети"'
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_1}_№_{number_prikaz_1}.pdf'
                    root_1[2][1][6].attrib['Kind'] = '01'
                else:
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_1}_№_{number_prikaz_1}.pdf'
                    root_1[2][1][5].attrib['Kind'] = '01'
                if not os.path.isdir(f'.//Готовые/ZoneToGKN_{str_guid_1}'):
                    os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}')
                os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}/Applied_files')
                tree.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//TerritoryToGKN_{str_guid}.xml',
                           encoding='utf-8', xml_declaration=True)
                tree_1.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//ZoneToGKN_{str_guid_1}.xml',
                             encoding='utf-8', xml_declaration=True)
                tree = ET.parse('Территория.xml')
                root = tree.getroot()
                tree_1 = ET.parse('Зона.xml')
                root_1 = tree_1.getroot()
    if kolvo_prikaz == 3:

        cod_documenta = input("Введите код первого документа(приказа): ")
        name_prikaz = input("Введите имя первого документа(приказа): ")
        number_prikaz = input("Введите номер первого документа(приказа): ")
        date_prikaz = input("Введите дату первого документа(приказа): ")
        publication = input("Подлежит ли документ(приказ) публикации? Введите да или нет: ")

        cod_documenta_1 = input("Введите код второго документа(приказа): ")
        name_prikaz_1 = input("Введите имя второго документа(приказа): ")
        number_prikaz_1 = input("Введите номер второго документа(приказа): ")
        date_prikaz_1 = input("Введите дату второго документа(приказа): ")
        publication_1 = input("Подлежит ли документ(приказ) публикации? Введите да или нет: ")

        cod_documenta_2 = input("Введите код третьего документа(приказа): ")
        name_prikaz_2 = input("Введите имя третьего документа(приказа): ")
        number_prikaz_2 = input("Введите номер третьего документа(приказа): ")
        date_prikaz_2 = input("Введите дату третьего документа(приказа): ")
        publication_2 = input("Подлежит ли документ(приказ) публикации? Введите да или нет: ")

        for i in range(len(total_sp)):
            ychastok = + i
            raion = total_sp[ychastok][2]
            zona = total_sp[ychastok][4][0]
            x = 3
            y = 4
            count = 1
            number_tochki = 6
            number_coord = 5
            # Условие, где 4 точки
            if len(total_sp[ychastok][12:]) == 1:
                for neighbor in root.iter(
                        '{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate'):
                    neighbor.attrib['X'] = total_sp[ychastok][x]
                    neighbor.attrib['Y'] = total_sp[ychastok][y]
                    x += 2
                    y += 2
                # Генерируем рандомный GUID
                guid = uuid.uuid4()
                # Конвертируем в строку
                str_guid = str(guid)
                root.attrib['GUID'] = str_guid
                root[0][1][0][1].text = "1047727043561"
                guid_1 = uuid.uuid4()
                str_guid_1 = str(guid_1)
                root[2][0].attrib["Name"] = f'МСК-50, зона {zona}'
                root_1.attrib['GUID'] = str_guid_1
                root_1[3][0][4][1][1].attrib['GUID'] = str_guid
                root_1[3][0][4][1][1].attrib['Name'] = f'TerritoryToGKN_{str_guid}.xml'
                root_1[0][1].text = f'ZoneToGKN_{str_guid_1}'
                root_1[2][0][0].text = cod_documenta
                root_1[0][3].text = date_prikaz
                root_1[2][0][1].text = name_prikaz
                root_1[2][0][2].text = number_prikaz
                root_1[2][0][3].text = date_prikaz
                if publication == "нет":
                    root_1[2][0][
                        5].tag = '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication'
                    root_1[2][0][6].attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
                else:
                    root_1[2][0].remove(root_1[2][0][5])
                    root_1[2][0][5].attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
                root_1[3][0][0].text = f'{total_sp[ychastok][2]}'
                root_1[3][0][2].text = f'{total_sp[ychastok][1]}'
                # Добавляем новый подэлемент (второй документ)
                ET.SubElement(root_1[2], '{urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8}Document')
                ET.SubElement(root_1[2], '{urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8}Document')
                # Добавляем в него еще подэлементы с новым текстом
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}CodeDocument').text = cod_documenta_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Name').text = name_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Number').text = number_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Date').text = date_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}IssueOrgan').text = 'Управление Росреестра по Московской области'
                if publication_1 == 'нет':
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication').text = 'Решение не подлежит публикации в связи с отсутствием необходимости опубликования таких решений в соответствии с Постановлением Правительства РФ от 21 августа 2019 г. N 1080 "Об охранных зонах пунктов государственной геодезической сети, государственной нивелирной сети и государственной гравиметрической сети"'
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_1}_№_{number_prikaz_1}.pdf'
                    root_1[2][1][6].attrib['Kind'] = '01'
                else:
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_1}_№_{number_prikaz_1}.pdf'
                    root_1[2][1][5].attrib['Kind'] = '01'

                # Добавляем третий документ
                # Добавляем в него еще подэлементы с новым текстом
                ET.SubElement(root_1[2][2],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}CodeDocument').text = cod_documenta_2
                ET.SubElement(root_1[2][2],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Name').text = name_prikaz_2
                ET.SubElement(root_1[2][2],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Number').text = number_prikaz_2
                ET.SubElement(root_1[2][2],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Date').text = date_prikaz_2
                ET.SubElement(root_1[2][2],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}IssueOrgan').text = 'Управление Росреестра по Московской области'
                if publication_1 == 'нет':
                    ET.SubElement(root_1[2][2],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication').text = 'Решение не подлежит публикации в связи с отсутствием необходимости опубликования таких решений в соответствии с Постановлением Правительства РФ от 21 августа 2019 г. N 1080 "Об охранных зонах пунктов государственной геодезической сети, государственной нивелирной сети и государственной гравиметрической сети"'
                    ET.SubElement(root_1[2][2],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_2}_№_{number_prikaz_2}.pdf'
                    root_1[2][2][6].attrib['Kind'] = '01'
                else:
                    ET.SubElement(root_1[2][2],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_2}_№_{number_prikaz_2}.pdf'
                    root_1[2][2][5].attrib['Kind'] = '01'

                if not os.path.isdir(f'.//Готовые/ZoneToGKN_{str_guid_1}'):
                    os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}')
                os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}/Applied_files')
                tree.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//TerritoryToGKN_{str_guid}.xml',
                           encoding='utf-8', xml_declaration=True)
                tree_1.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//ZoneToGKN_{str_guid_1}.xml',
                             encoding='utf-8', xml_declaration=True)
                tree = ET.parse('Территория.xml')
                root = tree.getroot()
                tree_1 = ET.parse('Зона.xml')
                root_1 = tree_1.getroot()
            # Условие, где  больше 4 точек
            if len(total_sp[ychastok][12:]) > 1:
                for neighbor_1 in root.iter(
                        '{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate'):
                    DeltaGeopoint = str(neighbor_1.attrib['DeltaGeopoint'])
                    GeopointOpred = str(neighbor_1.attrib['GeopointOpred'])
                    attrib_1 = {"TypeUnit": "Точка", "SuNmb": str(number_tochki)}
                    attrib_2 = {"X": str(total_sp[ychastok][x]), "Y": str(total_sp[ychastok][y]),
                                "NumGeopoint": str(number_tochki),
                                "DeltaGeopoint": DeltaGeopoint,
                                "GeopointOpred": GeopointOpred}
                    neighbor_1.attrib['X'] = total_sp[ychastok][x]
                    neighbor_1.attrib['Y'] = total_sp[ychastok][y]
                    x += 2
                    y += 2
                    count += 1
                    if count == 6:  # шестерка потому что, после пятой точки , становится 6 и тогда заходит в этот if.
                        root[1][0][4].attrib['SuNmb'] = '5'
                        neighbor_1.attrib['NumGeopoint'] = "5"
                        attrib_2['X'] = str(total_sp[ychastok][x])
                        attrib_2['Y'] = str(total_sp[ychastok][y])
                        count += 1
                        x += 2
                        y += 2
                    if count > 6:
                        konec = len(total_sp[i][13:])
                        while konec != 2:
                            ET.SubElement(root[1][0],
                                          "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}SpelementUnit",
                                          attrib_1)
                            ET.SubElement(root[1][0][number_coord],
                                          "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate ",
                                          attrib_2)
                            number_tochki += 1
                            attrib_1["SuNmb"] = str(number_tochki)
                            number_coord += 1
                            attrib_2['X'] = str(total_sp[ychastok][x])
                            attrib_2['Y'] = str(total_sp[ychastok][y])
                            attrib_2['NumGeopoint'] = str(number_tochki)
                            x += 2
                            y += 2
                            konec -= 2
                        attrib_1["SuNmb"] = str(1)
                        attrib_2['NumGeopoint'] = str(1)
                        ET.SubElement(root[1][0],
                                      "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}SpelementUnit",
                                      attrib_1)
                        ET.SubElement(root[1][0][number_coord],
                                      "{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}Ordinate ",
                                      attrib_2)
                # Генерируем рандомный GUID
                guid = uuid.uuid4()
                # Конвертируем в строку
                str_guid = str(guid)
                root.attrib['GUID'] = str_guid
                root[0][1][0][1].text = "1047727043561"
                guid_1 = uuid.uuid4()
                str_guid_1 = str(guid_1)
                root[2][0].attrib["Name"] = f'МСК-50, зона {zona}'
                root_1.attrib['GUID'] = str_guid_1
                root_1[3][0][4][1][1].attrib['GUID'] = str_guid
                root_1[3][0][4][1][1].attrib['Name'] = f'TerritoryToGKN_{str_guid}.xml'
                root_1[0][1].text = f'ZoneToGKN_{str_guid_1}'
                root_1[2][0][0].text = cod_documenta
                root_1[0][3].text = date_prikaz
                root_1[2][0][1].text = name_prikaz
                root_1[2][0][2].text = number_prikaz
                root_1[2][0][3].text = date_prikaz
                if publication == "нет":
                    root_1[2][0][
                        5].tag = '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication'
                    root_1[2][0][6].attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
                else:
                    root_1[2][0].remove(root_1[2][0][5])
                    root_1[2][0][5].attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz}_№_{number_prikaz}.pdf'
                root_1[3][0][0].text = f'{total_sp[ychastok][2]}'
                root_1[3][0][2].text = f'{total_sp[ychastok][1]}'
                # Добавляем новый подэлемент
                ET.SubElement(root_1[2], '{urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8}Document')
                ET.SubElement(root_1[2], '{urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8}Document')
                # Добавляем в него еще подэлементы с новым текстом
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}CodeDocument').text = cod_documenta_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Name').text = name_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Number').text = number_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Date').text = date_prikaz_1
                ET.SubElement(root_1[2][1],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}IssueOrgan').text = 'Управление Росреестра по Московской области'
                if publication_1 == 'нет':
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication').text = 'Решение не подлежит публикации в связи с отсутствием необходимости опубликования таких решений в соответствии с Постановлением Правительства РФ от 21 августа 2019 г. N 1080 "Об охранных зонах пунктов государственной геодезической сети, государственной нивелирной сети и государственной гравиметрической сети"'
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_1}_№_{number_prikaz_1}.pdf'
                    root_1[2][1][6].attrib['Kind'] = '01'
                else:
                    ET.SubElement(root_1[2][1],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_1}_№_{number_prikaz_1}.pdf'
                    root_1[2][1][5].attrib['Kind'] = '01'
                # Добавляем третий документ
                # Добавляем в него еще подэлементы с новым текстом
                ET.SubElement(root_1[2][2],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}CodeDocument').text = cod_documenta_2
                ET.SubElement(root_1[2][2],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Name').text = name_prikaz_2
                ET.SubElement(root_1[2][2],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Number').text = number_prikaz_2
                ET.SubElement(root_1[2][2],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}Date').text = date_prikaz_2
                ET.SubElement(root_1[2][2],
                              '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}IssueOrgan').text = 'Управление Росреестра по Московской области'
                if publication_1 == 'нет':
                    ET.SubElement(root_1[2][2],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}SourcePublication').text = 'Решение не подлежит публикации в связи с отсутствием необходимости опубликования таких решений в соответствии с Постановлением Правительства РФ от 21 августа 2019 г. N 1080 "Об охранных зонах пунктов государственной геодезической сети, государственной нивелирной сети и государственной гравиметрической сети"'
                    ET.SubElement(root_1[2][2],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_2}_№_{number_prikaz_2}.pdf'
                    root_1[2][2][6].attrib['Kind'] = '01'
                else:
                    ET.SubElement(root_1[2][2],
                                  '{urn://x-artefacts-rosreestr-ru/commons/complex-types/document-info/5.0.1}AppliedFile').attrib[
                        'Name'] = f'Applied_files\Приказ_Управления_от_{date_prikaz_2}_№_{number_prikaz_2}.pdf'
                    root_1[2][2][5].attrib['Kind'] = '01'

                if not os.path.isdir(f'.//Готовые/ZoneToGKN_{str_guid_1}'):
                    os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}')
                os.mkdir(f'.//Готовые/ZoneToGKN_{str_guid_1}/Applied_files')
                tree.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//TerritoryToGKN_{str_guid}.xml',
                           encoding='utf-8', xml_declaration=True)
                tree_1.write(f'.//Готовые/ZoneToGKN_{str_guid_1}//ZoneToGKN_{str_guid_1}.xml',
                             encoding='utf-8', xml_declaration=True)
                tree = ET.parse('Территория.xml')
                root = tree.getroot()
                tree_1 = ET.parse('Зона.xml')
                root_1 = tree_1.getroot()




# Теги внутренних элементов
# for new_ar in root_1[2][0]:
#     print(new_ar.tag)


# Смотрим какие элементы и подэлементы есть в xml
# for elm in root_1:
#     for sumelem in elm:
#         print(sumelem.tag)

# Добавить подэлемент
# ET.SubElement(root_1[2], '{urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8}Document')

# Удалить подэлемент
# root_1[1][0].remove(root_1[1][0][0])

# for xx in root.iter("{urn://x-artefacts-rosreestr-ru/commons/complex-types/entity-spatial/2.0.1}SpatialElement"):
#     print(xx[0].attrib)

"""Для запоминания"""
# Получается словарь (тег - ключ(name), attr - значение(название листа)).
# for child in root_1[2]:
#     print(child.tag, child.attrib)

# for b in root_1.iter("{urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8}Document"):
#     print(b[0].attrib)



# {urn://x-artefacts-rosreestr-ru/incoming/zone-to-gkn/5.0.8}Documents
"""worksheet {'name': 'грунтовые ГГС'}
worksheet {'name': 'на здании ГГС'}"""

#  Получить по индексу значение ячейки (.текст -  переводит в текст)
# print(root[0][0][1].text)
"""Район"""

# Заменяем определенную колонку на нужное значение и создаем новый файл хмл
# for raion in root.iter('Column3'):
#     raion.text = '50:15'

# Создаем столько файлов, сколько кортежей в общем списке, файлы создаются по названию поселка(в данном случае)
# for i in range(len(total_spisok)):
#     b =+ i
#     tree.write(f'{total_spisok[b][0]}.xml', encoding='utf-8')
input("Нажмите Enter для выхода")